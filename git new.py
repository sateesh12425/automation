import openpyxl
from openpyxl.styles import Font
from spellchecker import SpellChecker
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def check_spelling(text):
    if not text:
        return text

    spell = SpellChecker()
    words = text.split()
    corrected_words = []

    for word in words:
        candidates = spell.candidates(word)
        corrected_word = max(candidates, key=len) if candidates else word
        corrected_words.append(corrected_word)


    corrected_text = ' '.join(corrected_words)
    return corrected_text

def apply_styles_and_check_spelling(file_path, output_path, font_name='calibri', font_size=12):
    try:
        # Load the workbook and select the active worksheet
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):  # Check if the cell contains a string
                        # Spell check the cell value
                        corrected_text = check_spelling(cell.value)
                        cell.value = corrected_text

                        # Apply font style and size
                        cell.font = Font(name=font_name, size=font_size)
                        # cell.alignment = Alignment(horizontal='center', vertical='center')
        for cell in ws[1]:  # ws[1] accesses the first row
                cell.alignment = Alignment(horizontal='center', vertical='center')



        # Save the modified workbook
        wb.save(output_path)
        print(f"File saved as {output_path}")

    except FileNotFoundError:
        print(f"Error: The file {file_path} does not exist.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Example usage
input_file =r'C:\Users\nagar\OneDrive\Desktop\ravi_file.xlsx'
output_file = r'C:\Users\nagar\OneDrive\Desktop\corrected_file.xlsx'
apply_styles_and_check_spelling(input_file, output_file, font_name='calibri ', font_size=11)





